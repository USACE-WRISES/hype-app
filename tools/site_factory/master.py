"""Phase B: generate inputs_master.xlsx, the editable input of record.

One row per site on the SITES sheet. Derived cells (flow_cms, kh, kv) are live
Excel formulas so the user's tweaks propagate without re-running this script.
The driver reads this workbook, never the legacy Analysis workbook.

Usage:
  python tools/site_factory/master.py build          # full workbook, DESTROYS hand edits
  python tools/site_factory/master.py refresh-wells  # WELLS sheet only, SITES untouched,
                                                     # preserves name/include/screen_elev_ft
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))

HYPE_MODELS = Path(r"D:\Code\Work\hypoerheic-texas-sites\hype_models")
EXTRACTED = HYPE_MODELS / "_master" / "extracted"
OUT_XLSX = HYPE_MODELS / "_master" / "inputs_master.xlsx"

LL01096_TERRAIN = (r"D:\Code\Work\hypoerheic-texas-sites\Sites\LL01096\RAS_Post2021\Terrain"
                   r"\USGS_1m_FebMar2018_Revised_20260520_025.USGS_1m_FebMar2018_Revised_20260520_025.tif")

# (name, unit, contract mapping)
SITE_COLS = [
    ("site_id", "", "site.site_id"),
    ("river", "", "info"),
    ("lat", "deg", "site.outlet.lat"),
    ("lon", "deg", "site.outlet.lon"),
    ("include_flag", "", "gate"),
    ("model_flag", "", "gate"),
    ("confidence", "", "info"),
    ("status", "", "driver"),
    ("notes", "", "info"),
    ("geometry_source", "ras_import|auto", "geometry stage mode"),
    ("ras_prj_path", "path", "geometry stage input"),
    ("dem_source", "3dep|path", "terrain.dem_source"),
    ("dem_vertical_units", "m|ftUS", "terrain preprocessing"),
    ("flow_field_cfs", "cfs", "provenance only"),
    ("flow_field_provenance", "", "streamflow.provenance"),
    ("flow_model_cfs", "cfs", "provenance only"),
    ("flow_use_cfs", "cfs", "streamflow.value_cfs"),
    ("flow_cms", "m3/s", "streamflow.value_cms + ras payload flow_cms"),
    ("friction_slope", "ft/ft", "ras payload friction_slope"),
    ("manning_n", "", "ras payload manning_n"),
    ("k_site_ft_day", "ft/day", "provenance only"),
    ("kh_m_day", "m/day", "k.kh_m_day + params kh"),
    ("anisotropy_ratio", "", "k.anisotropy_ratio"),
    ("kv_m_day", "m/day", "k.kv_m_day + params kv"),
    ("porosity", "", "k.porosity + params porosity"),
    ("gradient_left", "m/m +=gaining", "gradients.left_controls"),
    ("gradient_right", "m/m +=gaining", "gradients.right_controls"),
    ("gradient_source", "", "gradients provenance"),
    ("ras_cell_m", "m", "ras payload cell_size_m"),
    ("gw_cell_m", "m", "grid.cell_size_x/y + params cell_size_x/y"),
    ("gw_mod_depth_m", "m", "grid.gw_mod_depth + params gw_mod_depth"),
    ("layer_thickness_m", "m", "grid.layer_thickness + params z"),
    ("hz_ppc", "", "hz params particles_per_cell"),
    ("sample_per_class", "", "hz params sample_per_class"),
    ("length_units", "literal", "params length_units"),
    ("time_units", "literal", "params time_units"),
    ("run_status", "", "driver-written"),
    ("run_id", "", "driver-written"),
    ("git_sha", "", "driver-written"),
    ("hash_geometry", "", "driver-written"),
    ("hash_terrain", "", "driver-written"),
    ("hash_streamflow", "", "driver-written"),
    ("hash_soil_k", "", "driver-written"),
    ("hash_gradients", "", "driver-written"),
    ("hash_grid", "", "driver-written"),
    ("failure_stage", "", "driver-written"),
    ("last_run_at", "", "driver-written"),
]

COLMAP_ROWS = [
    ("RULE", "Read the legacy workbook's SW Hydraulics sheet only. NEVER the OVERALL sheet: "
             "every OVERALL VLOOKUP into SW Hydraulics is off by one column (ColNumber+1), "
             "corrupting its HEC-RAS OUTPUT block (cols 47-52) and shifting every BRn label "
             "by one well. The single exception is the hand-entered Model (& Priority) flag."),
    ("RULE", "Key sites by FOLDER name, never by KMZ filename. LL01096's KMZ is literally "
             "named SS01208.kmz."),
    ("lat/lon", "Centroid of the site KMZ placemarks (Sites/<id>/GIS/*.kmz)."),
    ("flow_field_cfs", "SW Hydraulics 'Discharge (ft3/s)' with Method and date. Matches the "
                       "ADCP/FlowTracker .ctl files where present."),
    ("flow_model_cfs", "Final ordinate of the RAS unsteady flow file (<prj>.u01). "
                       "Calibration-adjusted, often differs from the field measurement."),
    ("flow_use_cfs", "DEFAULT = flow_model_cfs (reproduce the calibrated model). Change this "
                     "cell and re-run to model a different discharge."),
    ("friction_slope", "'Friction Slope' / 'Flow Hydrograph Slope' from the RAS .u01 "
                       "(identical US and DS in all 34 projects)."),
    ("k_site_ft_day", "SW Hydraulics 'Avg (ft/day)', the arithmetic mean of the 5 slug-test "
                      "well means (Hvorslev). Verified internally consistent."),
    ("kh_m_day", "= k_site_ft_day * 0.3048 (live formula)."),
    ("kv_m_day", "= kh_m_day / anisotropy_ratio (live formula). Anisotropy 10 is an "
                 "assumption, not a measurement."),
    ("porosity", "Assumed 0.30. Not measured anywhere in the source data."),
    ("gradient_left/right", "LL01096: uniform +0.001 (gaining) from Revised Gradient "
                            "Calculations.xlsx (Dec 2025), imposed value, both banks. "
                            "All other sites: 0.001 DEFAULT PENDING REVIEW. The older "
                            "per-well-pair gradients in SW Hydraulics cols 120-129 are an "
                            "order of magnitude larger and mixed-sign; review before batch."),
    ("geometry_source", "ras_import everywhere: all 34 sites carry a solved RAS 2D project "
                        "whose flow-area perimeter and US/DS BC lines are extractable from "
                        "<prj>.g01.hdf. No site needs NHD auto-delineation."),
    ("dem_source", "3dep default (native meters). LL01096 pilot uses the local May 20 2026 "
                   "flattened terrain export, ftUS vertical, converted to meters before "
                   "import. The exact calibrated terrain (Jun 25) exists only as raw 3DEP "
                   "plus unrasterized RAS-Mapper vector edits."),
    ("grid defaults", "ras_cell_m 3.0, gw_cell_m 2.0, gw_mod_depth_m 6.0, layer 0.5 "
                      "(12 layers). Retry ladder coarsens layer to 1.0 then cell to 3/4 m."),
    ("manning_n", "0.06 in every RAS project's 2D area attributes. Kept as a column so "
                  "per-site overrides are possible."),
    ("length_units", "Literal 'meters'. The engine's own default is feet and nothing "
                     "converts. Field names like *_cubic_ft hold metric values."),
    ("WELLS sheet", "One row per observation well harvested from the legacy GMS runs "
                    "(<site>.hob GMSCOMMENT XY + observed head, ftUS; names XY-matched "
                    "against TransObservation.csv / GIS Spikes points / Wells.shp because "
                    "the hob hedN order is scrambled). name, include and screen_elev_ft "
                    "are hand-editable and survive refresh-wells keyed by "
                    "(site_id, obs_name). NEVER hand-edit obs_name: it keys the app well "
                    "ids and the preserved edits. screen_elev_ft has no source anywhere; "
                    "fill it as field data becomes available and re-run the bundle stage."),
    ("WELLS include", "Yes rows with lat/lon become observation wells in the app project "
                      "at the bundle stage. The factory sets No when a well lands outside "
                      "3 km of the site centroid (CRS suspicion) or has no resolvable CRS."),
]

RUNS_COLS = ["run_id", "site_id", "started_at", "finished_at", "stages_run", "outcome",
             "failure_stage", "error", "git_sha", "input_hash"]

# WELLS sheet: one row per harvested GMS observation well (extract.py gw_wells).
# The first three data columns after the keys are HAND-EDITABLE and survive
# refresh-wells, keyed by (site_id, obs_name). Everything else is factory-owned.
WELLS_COLS = ["site_id", "obs_name", "name", "include", "screen_elev_ft",
              "lat", "lon", "obs_head_ft", "name_source", "dist_centroid_m",
              "source_hob", "note", "sw_head_ft", "gw_head_ft", "k_ft_s", "k_m_day"]
WELLS_EDITABLE = ("name", "include", "screen_elev_ft")


def _load_extracted_sites() -> list[dict]:
    sites = []
    for f in sorted(EXTRACTED.glob("*.json")):
        if f.name.startswith("_"):
            continue
        sites.append(json.loads(f.read_text(encoding="utf-8")))
    return sites


def _wells_rows_from_extracted(sites: list[dict]) -> list[dict]:
    """WELLS sheet rows: harvested gw_wells joined with the legacy per-well SW data.

    The legacy sw/gw/k columns join by the MATCHED name (BRn); a hand-renamed
    well keeps following the factory-matched name, which is factory-owned data.
    """
    rows = []
    for rec in sites:
        legacy = ((rec.get("sw_hydraulics") or {}).get("wells")) or {}
        for w in rec.get("gw_wells") or []:
            row = {c: None for c in WELLS_COLS}
            row.update({k: v for k, v in w.items() if k in row})
            row["site_id"] = rec["site_id"]
            leg = legacy.get(w.get("name")) or {}
            for k in ("sw_head_ft", "gw_head_ft", "k_ft_s", "k_m_day"):
                row[k] = leg.get(k)
            # Legacy sources can disagree per well: LL01096's GMS csv names vs
            # the SW Hydraulics head order (naming), CR08791/LL01869's local-
            # datum SW heads vs absolute model heads. Flag the mismatch
            # neutrally so the hand-editable name column gets a human look.
            gw, oh = row.get("gw_head_ft"), row.get("obs_head_ft")
            if (isinstance(gw, (int, float)) and isinstance(oh, (int, float))
                    and abs(gw - oh) > 0.2):
                conflict = (f"hob head {oh:.2f} vs SW Hydraulics {w.get('name')} "
                            f"gw head {gw:.2f} ft, review name and datum")
                row["note"] = f"{row['note']} + {conflict}" if row.get("note") else conflict
            rows.append(row)
    return rows


def _write_wells_sheet(ws, rows: list[dict], preserved: dict) -> None:
    """Fill a fresh WELLS worksheet.

    preserved maps (site_id, obs_name) -> {editable col: prior value}; a prior
    non-blank value in an editable column wins over the factory value, which is
    what makes refresh-wells safe to re-run after hand edits.
    """
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    white = Font(bold=True, size=10, color="FFFFFF")
    fill_head = PatternFill("solid", fgColor="1F3864")
    fill_edit_head = PatternFill("solid", fgColor="375623")
    fill_factory = PatternFill("solid", fgColor="EEEEEE")

    for j, h in enumerate(WELLS_COLS, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = white
        c.fill = fill_edit_head if h in WELLS_EDITABLE else fill_head
        width = 34 if h == "source_hob" else max(10, min(len(h) + 3, 22))
        ws.column_dimensions[get_column_letter(j)].width = width
    r = 2
    for row in rows:
        keep = preserved.get((row.get("site_id"), row.get("obs_name"))) or {}
        for j, h in enumerate(WELLS_COLS, 1):
            val = row.get(h)
            if h in WELLS_EDITABLE:
                pv = keep.get(h)
                if pv is not None and (not isinstance(pv, str) or pv.strip()):
                    val = pv
            c = ws.cell(row=r, column=j, value=val)
            if h not in WELLS_EDITABLE:
                c.fill = fill_factory
        r += 1
    ws.freeze_panes = "C2"


def build():
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sites = _load_extracted_sites()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SITES"

    hdr_font = Font(bold=True, size=10)
    meta_font = Font(size=8, italic=True, color="666666")
    fill_head = PatternFill("solid", fgColor="1F3864")
    fill_driver = PatternFill("solid", fgColor="EEEEEE")
    white = Font(bold=True, size=10, color="FFFFFF")

    for j, (name, unit, contract) in enumerate(SITE_COLS, 1):
        c = ws.cell(row=1, column=j, value=name)
        c.font = white
        c.fill = fill_head
        ws.cell(row=2, column=j, value=unit).font = meta_font
        ws.cell(row=3, column=j, value=contract).font = meta_font
        ws.column_dimensions[get_column_letter(j)].width = max(11, min(len(name) + 2, 22))

    col_idx = {name: j for j, (name, _, _) in enumerate(SITE_COLS, 1)}

    def put(r, name, value):
        ws.cell(row=r, column=col_idx[name], value=value)

    r = 4
    for rec in sites:
        sid = rec["site_id"]
        sw = rec.get("sw_hydraulics") or {}
        coords = rec.get("coords") or {}
        ras = (rec.get("ras_details") or [{}])[0]
        flow_model = None
        fslope = None
        for fl in ras.get("flows", []):
            flow_model = fl.get("flow_final_cfs") or flow_model
            fslope = fl.get("ds_friction_slope") or fl.get("us_flow_slope") or fslope
        n_val = None
        for g in ras.get("geometries", []):
            if g.get("mannings_n"):
                n_val = round(g["mannings_n"], 3)
                break
        gwb = rec.get("gradient_workbook") or {}
        grad = gwb.get("uniform_gradient")
        grad_src = (Path(gwb["file"]).name + " (imposed, back-calculated heads)"
                    if grad is not None and gwb.get("file") else "default 0.001 pending review")
        is_pilot = sid == "LL01096"

        put(r, "site_id", sid)
        put(r, "river", sw.get("river"))
        put(r, "lat", coords.get("lat"))
        put(r, "lon", coords.get("lon"))
        put(r, "include_flag", sw.get("include"))
        put(r, "model_flag", rec.get("model_flag"))
        put(r, "confidence", sw.get("confidence"))
        put(r, "status", "ready_pilot" if is_pilot else "needs_review")
        put(r, "notes", sw.get("sw_notes"))
        put(r, "geometry_source", "ras_import")
        put(r, "ras_prj_path", (rec.get("ras_projects") or [None])[0])
        put(r, "dem_source", LL01096_TERRAIN if is_pilot else "3dep")
        put(r, "dem_vertical_units", "ftUS" if is_pilot else "m")
        put(r, "flow_field_cfs", sw.get("discharge_cfs"))
        put(r, "flow_field_provenance",
            f"{sw.get('discharge_method') or '?'} {sw.get('discharge_date') or ''}".strip())
        put(r, "flow_model_cfs", flow_model)
        put(r, "flow_use_cfs", flow_model if flow_model is not None else sw.get("discharge_cfs"))
        fc = get_column_letter(col_idx["flow_use_cfs"])
        ws.cell(row=r, column=col_idx["flow_cms"], value=f"={fc}{r}*0.028316846592")
        put(r, "friction_slope", fslope)
        put(r, "manning_n", n_val or 0.06)
        put(r, "k_site_ft_day", sw.get("k_avg_ft_day"))
        kc = get_column_letter(col_idx["k_site_ft_day"])
        ws.cell(row=r, column=col_idx["kh_m_day"], value=f"={kc}{r}*0.3048")
        put(r, "anisotropy_ratio", 10.0)
        khc = get_column_letter(col_idx["kh_m_day"])
        ac = get_column_letter(col_idx["anisotropy_ratio"])
        ws.cell(row=r, column=col_idx["kv_m_day"], value=f"={khc}{r}/{ac}{r}")
        put(r, "porosity", 0.30)
        put(r, "gradient_left", grad if grad is not None else 0.001)
        put(r, "gradient_right", grad if grad is not None else 0.001)
        put(r, "gradient_source", grad_src)
        put(r, "ras_cell_m", 3.0)
        put(r, "gw_cell_m", 2.0)
        put(r, "gw_mod_depth_m", 6.0)
        put(r, "layer_thickness_m", 0.5)
        put(r, "hz_ppc", 1)
        put(r, "sample_per_class", 300)
        put(r, "length_units", "meters")
        put(r, "time_units", "days")
        for name in ("run_status", "run_id", "git_sha", "hash_geometry", "hash_terrain",
                     "hash_streamflow", "hash_soil_k", "hash_gradients", "hash_grid",
                     "failure_stage", "last_run_at"):
            ws.cell(row=r, column=col_idx[name]).fill = fill_driver
        r += 1

    ws.freeze_panes = "B4"

    # WELLS
    wells_rows = _wells_rows_from_extracted(sites)
    _write_wells_sheet(wb.create_sheet("WELLS"), wells_rows, {})
    wr = len(wells_rows) + 2

    # COLMAP
    wc = wb.create_sheet("COLMAP")
    wc.cell(row=1, column=1, value="item").font = hdr_font
    wc.cell(row=1, column=2, value="source and rules").font = hdr_font
    wc.column_dimensions["A"].width = 24
    wc.column_dimensions["B"].width = 130
    for i, (k, v) in enumerate(COLMAP_ROWS, 2):
        wc.cell(row=i, column=1, value=k).font = hdr_font if k == "RULE" else None
        cell = wc.cell(row=i, column=2, value=v)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # RUNS
    wr_ = wb.create_sheet("RUNS")
    for j, h in enumerate(RUNS_COLS, 1):
        c = wr_.cell(row=1, column=j, value=h)
        c.font = white
        c.fill = fill_head

    meta = wb.create_sheet("README")
    meta.cell(row=1, column=1, value="inputs_master.xlsx, generated "
              + datetime.now(timezone.utc).isoformat(timespec="seconds"))
    meta.cell(row=2, column=1, value="SITES is the input of record. Edit values, then re-run the "
              "driver: only the stages whose dependency-group hash changed will re-execute.")
    meta.cell(row=3, column=1, value="Driver-written columns are gray. Do not hand-edit them.")

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX} ({r - 4} site rows, {wr - 2} well rows)")


def read_sites() -> dict:
    """Driver-side reader: {site_id: {col: value}} with formulas evaluated by openpyxl data_only.

    Note data_only returns cached values, which exist only after the workbook has
    been opened and saved by Excel once. Until then, fall back to computing the
    two live formulas ourselves.
    """
    import openpyxl

    wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
    ws = wb["SITES"]
    names = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        rec = dict(zip(names, row))
        sid = rec.get("site_id")
        if not sid:
            continue
        if rec.get("flow_cms") is None and rec.get("flow_use_cfs") is not None:
            rec["flow_cms"] = rec["flow_use_cfs"] * 0.028316846592
        if rec.get("kh_m_day") is None and rec.get("k_site_ft_day") is not None:
            rec["kh_m_day"] = rec["k_site_ft_day"] * 0.3048
        if rec.get("kv_m_day") is None and rec.get("kh_m_day") is not None:
            rec["kv_m_day"] = rec["kh_m_day"] / (rec.get("anisotropy_ratio") or 10.0)
        out[str(sid)] = rec
    wb.close()
    return out


def refresh_wells():
    """Rebuild ONLY the WELLS sheet of the existing workbook, in place.

    SITES and every other sheet are untouched, so hand edits there survive.
    Hand edits in the WELLS editable columns survive too, keyed by
    (site_id, obs_name). openpyxl re-save drops Excel's cached formula values
    on SITES; read_sites() computes those three columns itself until Excel
    next opens and saves the workbook.
    """
    import openpyxl

    if not OUT_XLSX.exists():
        raise SystemExit(f"{OUT_XLSX} missing, run build first")
    rows = _wells_rows_from_extracted(_load_extracted_sites())
    wb = openpyxl.load_workbook(OUT_XLSX)
    preserved = {}
    if "WELLS" in wb.sheetnames:
        ws = wb["WELLS"]
        hdr = [c.value for c in ws[1]]
        if "obs_name" in hdr:
            for vals in ws.iter_rows(min_row=2, values_only=True):
                rec = dict(zip(hdr, vals))
                key = (rec.get("site_id"), rec.get("obs_name"))
                if all(key):
                    preserved[key] = {k: rec.get(k) for k in WELLS_EDITABLE}
        idx = wb.sheetnames.index("WELLS")
        del wb["WELLS"]
        ws_new = wb.create_sheet("WELLS", idx)
    else:
        ws_new = wb.create_sheet("WELLS")
    _write_wells_sheet(ws_new, rows, preserved)
    wb.save(OUT_XLSX)
    kept = sum(1 for v in preserved.values() if any(x is not None for x in v.values()))
    print(f"refreshed WELLS in {OUT_XLSX}: {len(rows)} rows, {kept} prior rows carried edits")
    print("note: SITES cached formula values reset until Excel opens and saves the workbook once")


def read_wells() -> dict:
    """Driver-side reader: {site_id: [WELLS row dicts]}.

    Returns {} when the sheet is absent or still the legacy (pre-obs_name)
    format, so the driver degrades to no injected wells instead of crashing.
    """
    import openpyxl

    if not OUT_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(OUT_XLSX, data_only=True)
    if "WELLS" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["WELLS"]
    hdr = [c.value for c in ws[1]]
    if "obs_name" not in hdr:
        wb.close()
        return {}
    out = {}
    for vals in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(hdr, vals))
        sid = rec.get("site_id")
        if not sid or not rec.get("obs_name"):
            continue
        out.setdefault(str(sid), []).append(rec)
    wb.close()
    return out


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else None
    if cmd == "build":
        build()
    elif cmd == "refresh-wells":
        refresh_wells()
    else:
        print(__doc__)
