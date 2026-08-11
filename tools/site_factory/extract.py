"""Phase A extraction: read every Texas site's source data into one JSON per site.

Sources, in order of trust:
  1. SW Hydraulics sheet of the master workbook (row per site, labels on row 3).
     NEVER the OVERALL sheet: every OVERALL VLOOKUP into SW Hydraulics is off by
     one column, which corrupts its whole HEC-RAS OUTPUT block and shifts every
     BRn label by one well. OVERALL is read for exactly one thing, the
     "Model (& Priority)" flag, which is a hand-entered column.
  2. Per-site files: KMZ (coordinates), RAS project (flow, slope, geometry),
     gradient-calculation workbook, discharge .ctl, slug-test workbook path.

A self-test pins the SW Hydraulics column map against three independently
verified LL01096 values before any other site is read. If the map drifts, we
abort rather than mis-extract 34 sites.

Usage:
  python tools/site_factory/extract.py                 # all sites
  python tools/site_factory/extract.py --site LL01096  # one site
"""
from __future__ import annotations

import argparse
import io
import json
import re
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO))

SITES_ROOT = Path(r"D:\Code\Work\hypoerheic-texas-sites\Sites")
MASTER_XLSX = Path(r"D:\Code\Work\hypoerheic-texas-sites\Analysis\2025-01-26_SRW Summary Data_Updated.xlsx")
OUT_DIR = Path(r"D:\Code\Work\hypoerheic-texas-sites\hype_models\_master\extracted")

FT_PER_M = 3.280839895
FT_S_TO_M_DAY = 86400 * 0.3048

# LL01096 pins, each verified by hand this week (SW Hydraulics row 18).
SELFTEST = {
    "site": "LL01096",
    "discharge_cfs": 124.38543634,
    "k_avg_ft_day": 685.003675463424,
    "us_ds_wss": 0.008881249104712793,
}


# ---------------------------------------------------------------- workbook
def _cell(row, idx1):
    """1-based column access into a values-only row tuple."""
    v = row[idx1 - 1] if idx1 - 1 < len(row) else None
    if isinstance(v, str) and v.strip() in ("#N/A", "#VALUE!", "#DIV/0!", "#REF!", ""):
        return None
    return v


def load_sw_hydraulics():
    """Return {site_id: fields} from the SW Hydraulics sheet.

    The column map is located by row-3 labels (with the repeating BR1..BR5
    groups disambiguated by the row-2 group header at each group start), then
    pinned by the LL01096 self-test.
    """
    import openpyxl

    wb = openpyxl.load_workbook(MASTER_XLSX, data_only=True, read_only=True)
    ws = wb["SW Hydraulics"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=150, values_only=True))
    r2, r3 = rows[1], rows[2]

    def find(label, after=0):
        for i in range(after, 150):
            if r3[i] is not None and str(r3[i]).strip() == label:
                return i + 1
        raise KeyError(label)

    def group_start(label_frag):
        for i in range(150):
            if r2[i] is not None and label_frag in str(r2[i]):
                return i + 1
        raise KeyError(label_frag)

    C = {
        "river": find("River"),
        "site": find("Site ID"),
        "projection": find("Projection"),
        "sw_notes": find("Notes", after=find("Projection")),
        "confidence": find("Confidence"),
        "include": find("Include Site?"),
        "ds_bc_slope": find("DS BC Slope (ft/ft)"),
        "q_year": find("Year"),
        "q_month": find("Month"),
        "q_day": find("Day"),
        "discharge_cfs": find("Discharge (ft³/s)"),
        "method": find("Method"),
        "terrain_used": find("TERRAIN USED?"),
        "wse_slope": find("WSE Slope (ft/ft)"),
        "us_ds_wse": find("US-DS WSE (ft)"),
        "us_ds_wss": find("US-DS WSS (ft/ft)"),
        "k_avg_ft_s": find("Avg", after=group_start("K (ft/s)")),
        "k_avg_ft_day": find("Avg (ft/day)"),
        "stream_width_ft": find("Stream Width (ft)"),
    }
    C["wse_br1"] = group_start("Outside Spike WSE")
    C["gw_br1"] = group_start("GW Elev (inside well)")
    C["k_br1"] = group_start("K (ft/s)")
    C["grad_br1"] = group_start("GRADIENTS (NEGATIVE MEANS FL")

    out = {}
    for row in rows[3:]:
        sid = _cell(row, C["site"])
        if not sid or not isinstance(sid, str):
            continue
        sid = sid.strip()
        rec = {
            "river": _cell(row, C["river"]),
            "projection": _cell(row, C["projection"]),
            "sw_notes": _cell(row, C["sw_notes"]),
            "confidence": _cell(row, C["confidence"]),
            "include": _cell(row, C["include"]),
            "ds_bc_slope_ft_ft": _cell(row, C["ds_bc_slope"]),
            "discharge_cfs": _cell(row, C["discharge_cfs"]),
            "discharge_method": _cell(row, C["method"]),
            "discharge_date": "-".join(
                str(int(v)) if isinstance(v, (int, float)) else str(v)
                for v in (_cell(row, C["q_year"]), _cell(row, C["q_month"]), _cell(row, C["q_day"]))
                if v is not None
            ) or None,
            "terrain_used": _cell(row, C["terrain_used"]),
            "wse_slope_ft_ft": _cell(row, C["wse_slope"]),
            "us_ds_wse_ft": _cell(row, C["us_ds_wse"]),
            "us_ds_wss_ft_ft": _cell(row, C["us_ds_wss"]),
            "k_avg_ft_s": _cell(row, C["k_avg_ft_s"]),
            "k_avg_ft_day": _cell(row, C["k_avg_ft_day"]),
            "stream_width_ft": _cell(row, C["stream_width_ft"]),
            "wells": {},
        }
        for w in range(5):
            name = f"BR{w + 1}"
            k_ft_s = _cell(row, C["k_br1"] + w)
            rec["wells"][name] = {
                "sw_head_ft": _cell(row, C["wse_br1"] + w),
                "gw_head_ft": _cell(row, C["gw_br1"] + w),
                "k_ft_s": k_ft_s,
                "k_m_day": (k_ft_s * FT_S_TO_M_DAY) if isinstance(k_ft_s, (int, float)) else None,
            }
        pair_names = ["BR1/BR2", "BR1/BR3", "BR1/BR4", "BR1/BR5", "BR2/BR3",
                      "BR4/BR3", "BR5/BR3", "BR2/BR4", "BR5/BR2", "BR5/BR4"]
        rec["gw_gradients"] = {
            nm: _cell(row, C["grad_br1"] + i) for i, nm in enumerate(pair_names)
        }
        out[sid] = rec
    wb.close()

    pin = out.get(SELFTEST["site"])
    if pin is None:
        raise SystemExit("SELF-TEST FAILED: LL01096 row not found in SW Hydraulics")
    checks = [
        ("discharge_cfs", pin["discharge_cfs"], SELFTEST["discharge_cfs"]),
        ("k_avg_ft_day", pin["k_avg_ft_day"], SELFTEST["k_avg_ft_day"]),
        ("us_ds_wss", pin["us_ds_wss_ft_ft"], SELFTEST["us_ds_wss"]),
    ]
    for name, got, want in checks:
        if got is None or abs(got - want) > abs(want) * 1e-9 + 1e-12:
            raise SystemExit(f"SELF-TEST FAILED on {name}: got {got!r}, want {want!r}. "
                             "Column map has drifted, refusing to extract.")
    return out


def load_model_flags():
    """OVERALL is corrupt except hand-entered columns. Read only Model (& Priority)."""
    import openpyxl

    wb = openpyxl.load_workbook(MASTER_XLSX, data_only=True, read_only=True)
    ws = wb["OVERALL"]
    rows = list(ws.iter_rows(min_row=1, max_row=45, max_col=140, values_only=True))
    labels = rows[4]  # row 5
    col = None
    for i, v in enumerate(labels):
        if v is not None and "Model" in str(v) and "Priority" in str(v):
            col = i + 1
            break
    if col is None:
        return {}
    site_col = None
    for i, v in enumerate(labels):
        if v is not None and str(v).strip() in ("Site", "Site ID", "Site Name"):
            site_col = i + 1
            break
    out = {}
    for row in rows[5:43]:
        sid = None
        if site_col:
            sid = _cell(row, site_col)
        if not sid:
            for v in row[:6]:
                if isinstance(v, str) and re.fullmatch(r"[A-Z]{2}\d{5}", v.strip()):
                    sid = v.strip()
                    break
        if not sid:
            continue
        out[str(sid).strip()] = _cell(row, col)
    wb.close()
    return out


# ---------------------------------------------------------------- site files
def kmz_centroid(site_dir: Path):
    """Centroid of all coordinates in the site's first KMZ. Keyed by folder, never filename."""
    kmzs = sorted(site_dir.rglob("*.kmz"))
    if not kmzs:
        return None
    kmz = kmzs[0]
    try:
        with zipfile.ZipFile(kmz) as z:
            kml_names = [n for n in z.namelist() if n.lower().endswith(".kml")]
            if not kml_names:
                return None
            text = z.read(kml_names[0]).decode("utf-8", "replace")
    except Exception as e:  # noqa: BLE001
        return {"kmz": str(kmz), "error": str(e)}
    pts = []
    for m in re.finditer(r"<coordinates>(.*?)</coordinates>", text, re.S):
        for tok in m.group(1).split():
            parts = tok.split(",")
            if len(parts) >= 2:
                try:
                    lon, lat = float(parts[0]), float(parts[1])
                except ValueError:
                    continue
                if -107 < lon < -93 and 25 < lat < 37:
                    pts.append((lon, lat))
    if not pts:
        return {"kmz": str(kmz), "error": "no coordinates parsed"}
    lon = sum(p[0] for p in pts) / len(pts)
    lat = sum(p[1] for p in pts) / len(pts)
    return {
        "kmz": str(kmz),
        "kmz_name_matches_site": kmz.stem.upper() == site_dir.name.upper(),
        "n_points": len(pts),
        "lat": round(lat, 6),
        "lon": round(lon, 6),
    }


def is_ras_prj(path: Path) -> bool:
    try:
        head = path.read_bytes()[:400].decode("ascii", "replace")
    except Exception:  # noqa: BLE001
        return False
    return "Proj Title=" in head


def parse_u_file(path: Path):
    """Flow ordinates and slopes out of a RAS unsteady flow file (plain text)."""
    try:
        text = path.read_text(errors="replace")
    except Exception as e:  # noqa: BLE001
        return {"error": str(e)}
    out = {"file": path.name}
    m = re.search(r"Flow Hydrograph=\s*(\d+)\s*\n((?:[ \d.\-+eE]+\n)+)", text)
    if m:
        vals = [float(v) for v in m.group(2).split()]
        out["flow_ordinates_cfs"] = vals
        out["flow_final_cfs"] = vals[-1] if vals else None
    m = re.search(r"Flow Hydrograph Slope=\s*([\d.eE\-+]+)", text)
    if m:
        out["us_flow_slope"] = float(m.group(1))
    m = re.search(r"Friction Slope=\s*([\d.eE\-+]+)", text)
    if m:
        out["ds_friction_slope"] = float(m.group(1))
    return out


def inspect_ras_project(prj: Path):
    """Project structure, flow, slope, and g0N.hdf geometry availability."""
    text = prj.read_text(errors="replace")
    rec = {
        "prj": str(prj),
        "title": (re.search(r"Proj Title=(.*)", text) or [None, None])[1],
        "current_plan": (re.search(r"Current Plan=(\w+)", text) or [None, None])[1],
        "units": "US Customary" if "English Units" in text else ("SI" if "SI Units" in text else None),
        "geom_files": re.findall(r"Geom File=(\w+)", text),
        "plan_files": re.findall(r"Plan File=(\w+)", text),
        "unsteady_files": re.findall(r"Unsteady File=(\w+)", text),
    }
    base = prj.with_suffix("")
    for u in rec["unsteady_files"]:
        up = base.parent / f"{base.name}.{u}"
        if up.exists():
            rec.setdefault("flows", []).append(parse_u_file(up))
    rec["rasmap"] = (base.parent / f"{base.name}.rasmap").exists()
    for g in rec["geom_files"]:
        ghdf = base.parent / f"{base.name}.{g}.hdf"
        if not ghdf.exists():
            continue
        grec = {"file": ghdf.name, "size": ghdf.stat().st_size}
        try:
            import h5py

            with h5py.File(ghdf, "r") as f:
                geo = f.get("Geometry")
                if geo is not None:
                    fa = geo.get("2D Flow Areas")
                    grec["has_2d_area"] = fa is not None
                    if fa is not None:
                        names = [k for k in fa.keys() if isinstance(fa.get(k), h5py.Group)]
                        grec["flow_areas"] = names
                        for nm in names:
                            per = fa[nm].get("Perimeter")
                            if per is not None:
                                grec["perimeter_vertices"] = int(per.shape[0])
                                break
                        attrs = fa.get("Attributes")
                        if attrs is not None:
                            try:
                                a0 = attrs[0]
                                grec["cell_spacing_ft"] = [float(a0["Spacing dx"]), float(a0["Spacing dy"])]
                                grec["mannings_n"] = float(a0["Mann"])
                            except Exception:  # noqa: BLE001
                                pass
                    bc = geo.get("Boundary Condition Lines")
                    if bc is not None and bc.get("Attributes") is not None:
                        grec["bc_lines"] = [
                            (r[0].decode() if isinstance(r[0], bytes) else str(r[0]))
                            for r in bc["Attributes"][:]
                        ]
                    grec["has_centerline"] = geo.get("River Centerlines") is not None
        except Exception as e:  # noqa: BLE001
            grec["h5_error"] = str(e)
        rec.setdefault("geometries", []).append(grec)
    return rec


def gradient_workbook(site_dir: Path):
    """Read a Revised Gradient Calculations workbook if present (LL01096 pattern)."""
    cands = [p for p in site_dir.rglob("*.xlsx") if "gradient" in p.name.lower()]
    if not cands:
        return None
    path = cands[0]
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=1, max_row=20, max_col=8, values_only=True))
        wb.close()
        hdr = [str(v) if v is not None else "" for v in rows[0]]
        if "Gradient" not in hdr:
            return {"file": str(path), "pattern": "unrecognized"}
        gcol = hdr.index("Gradient")
        vals = [r[gcol] for r in rows[1:] if r[gcol] is not None and isinstance(r[gcol], (int, float))]
        uniform = len(set(round(v, 9) for v in vals)) == 1 if vals else None
        return {
            "file": str(path),
            "pattern": "wse_backcalc",
            "n_points": len(vals),
            "gradients": vals[:12],
            "uniform_gradient": vals[0] if uniform and vals else None,
        }
    except Exception as e:  # noqa: BLE001
        return {"file": str(path), "error": str(e)}


def discharge_ctl(site_dir: Path):
    for ctl in site_dir.rglob("*.ctl"):
        try:
            text = ctl.read_text(errors="replace")
        except Exception:  # noqa: BLE001
            continue
        m = re.search(r"Total_Discharge\s+([\d.eE\-+]+)", text)
        if m:
            return {"file": str(ctl), "total_discharge_cms": float(m.group(1))}
    return None


def inventory(site_dir: Path):
    """Targeted globs only. The tree is 22 GB and layouts drift, so glob, never hardcode."""
    ras_prjs = [p for p in site_dir.rglob("*.prj") if is_ras_prj(p)]
    gprs = sorted(site_dir.rglob("*.gpr"))
    tifs = sorted(site_dir.rglob("*.tif"), key=lambda p: -p.stat().st_size)
    return {
        "ras_projects": [str(p) for p in ras_prjs],
        "gms_projects": [str(p) for p in gprs],
        "dem_candidates": [
            {"path": str(p), "mb": round(p.stat().st_size / 1e6, 1)} for p in tifs[:6]
        ],
        "slug_workbooks": [str(p) for p in site_dir.rglob("*slugtest*.xlsx")],
        "wells_shp": [str(p) for p in site_dir.rglob("Wells*.shp")],
        "profile_lines_shp": [str(p) for p in site_dir.rglob("Profile Lines.shp")],
    }


# ---------------------------------------------------------------- main
def extract_site(site_dir: Path, sw_row, model_flag):
    from tools.site_factory import hob

    rec = {
        "site_id": site_dir.name,
        "extracted_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "coords": kmz_centroid(site_dir),
        "sw_hydraulics": sw_row,
        "model_flag": model_flag,
        "gradient_workbook": gradient_workbook(site_dir),
        "discharge_ctl": discharge_ctl(site_dir),
    }
    rec.update(inventory(site_dir))
    rec["ras_details"] = [inspect_ras_project(Path(p)) for p in rec["ras_projects"]]
    coords = rec["coords"] if isinstance(rec["coords"], dict) else {}
    rec["gw_wells"] = hob.harvest_site_wells(
        site_dir, site_dir.name, coords.get("lat"), coords.get("lon"))
    return rec


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--site", help="extract one site only")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    sw = load_sw_hydraulics()
    print(f"SW Hydraulics: {len(sw)} site rows, self-test PASSED")
    flags = load_model_flags()
    print(f"OVERALL model flags: {len(flags)} rows")

    dirs = sorted(d for d in SITES_ROOT.iterdir() if d.is_dir())
    if args.site:
        dirs = [d for d in dirs if d.name.upper() == args.site.upper()]
        if not dirs:
            raise SystemExit(f"no folder for {args.site}")

    summary = []
    for d in dirs:
        rec = extract_site(d, sw.get(d.name), flags.get(d.name))
        out = OUT_DIR / f"{d.name}.json"
        out.write_text(json.dumps(rec, indent=2, default=str), encoding="utf-8")
        n_ras = len(rec["ras_projects"])
        flow = None
        for rd in rec["ras_details"]:
            for fl in rd.get("flows", []):
                flow = fl.get("flow_final_cfs") or flow
        summary.append({
            "site": d.name,
            "coords": bool(rec["coords"] and rec["coords"].get("lat")),
            "sw_row": rec["sw_hydraulics"] is not None,
            "model_flag": rec["model_flag"],
            "ras": n_ras,
            "ras_flow_cfs": flow,
            "gms": len(rec["gms_projects"]),
            "wells": len(rec["gw_wells"]),
            "wells_named": sum(1 for w in rec["gw_wells"]
                               if w.get("name_source") not in (None, "hob")),
            "field_q_cfs": (rec["sw_hydraulics"] or {}).get("discharge_cfs"),
            "k_m_day": round((rec["sw_hydraulics"] or {}).get("k_avg_ft_day") * 0.3048, 2)
            if (rec["sw_hydraulics"] or {}).get("k_avg_ft_day") else None,
            "geometry_mode": "ras_import" if any(
                g.get("has_2d_area") for rd in rec["ras_details"] for g in rd.get("geometries", [])
            ) else "auto",
        })
        print(f"  {d.name}: ras={n_ras} gms={len(summary[-1]) and summary[-1]['gms']} "
              f"mode={summary[-1]['geometry_mode']} flow={flow} "
              f"wells={summary[-1]['wells']} named={summary[-1]['wells_named']}")

    (OUT_DIR / "_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(f"\nwrote {len(summary)} site JSONs -> {OUT_DIR}")


if __name__ == "__main__":
    main()
