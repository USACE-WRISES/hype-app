"""Unit tests for the inputs_master.xlsx wells pipeline (master.py).

The load-bearing property: `refresh-wells` rewrites ONLY the WELLS sheet, so
hand edits on SITES (and the three live formula cells) survive, and the WELLS
editable columns (name, include, screen_elev_ft) survive keyed by
(site_id, obs_name).
"""
from __future__ import annotations

import json

import openpyxl
import pytest

from tools.site_factory import master


def _site_json(sid="AA00001", obs_head=1311.6, extra_well=False):
    wells = [{
        "obs_name": "hed1", "name": "BR1", "name_source": "spikes",
        "lat": 30.0, "lon": -99.0, "obs_head_ft": obs_head,
        "dist_centroid_m": 50.0, "include": "Yes", "note": None,
        "source_hob": "GMS/A_MODFLOW/A.hob"}]
    if extra_well:
        wells.append({
            "obs_name": "hed2", "name": "BR2", "name_source": "hob",
            "lat": 30.001, "lon": -99.001, "obs_head_ft": 1310.0,
            "dist_centroid_m": 60.0, "include": "Yes", "note": None,
            "source_hob": "GMS/A_MODFLOW/A.hob"})
    return {
        "site_id": sid,
        "coords": {"lat": 30.0, "lon": -99.0},
        "sw_hydraulics": {
            "river": "Test", "include": "Yes", "confidence": "High", "sw_notes": None,
            "discharge_cfs": 10.0, "discharge_method": "ADCP", "discharge_date": "2021-1-1",
            "k_avg_ft_day": 100.0, "projection": None,
            "wells": {"BR1": {"sw_head_ft": 1315.0, "gw_head_ft": 1311.5,
                              "k_ft_s": 0.001, "k_m_day": 26.3}},
        },
        "model_flag": "Yes",
        "gradient_workbook": None, "discharge_ctl": None,
        "ras_projects": ["C:/fake/site.prj"], "gms_projects": [],
        "dem_candidates": [], "slug_workbooks": [], "wells_shp": [],
        "profile_lines_shp": [], "ras_details": [],
        "gw_wells": wells,
    }


@pytest.fixture()
def workbook_env(tmp_path, monkeypatch):
    extracted = tmp_path / "extracted"
    extracted.mkdir()
    (extracted / "AA00001.json").write_text(json.dumps(_site_json()), encoding="utf-8")
    out = tmp_path / "inputs_master.xlsx"
    monkeypatch.setattr(master, "EXTRACTED", extracted)
    monkeypatch.setattr(master, "OUT_XLSX", out)
    master.build()
    return extracted, out


def _col(ws, header):
    return [c.value for c in ws[1]].index(header) + 1


def test_build_writes_enriched_wells_sheet(workbook_env):
    _, out = workbook_env
    wb = openpyxl.load_workbook(out)
    ws = wb["WELLS"]
    hdr = [c.value for c in ws[1]]
    assert hdr == list(master.WELLS_COLS)
    row = [c.value for c in ws[2]]
    rec = dict(zip(hdr, row))
    assert rec["site_id"] == "AA00001"
    assert rec["obs_name"] == "hed1"
    assert rec["name"] == "BR1"
    assert rec["obs_head_ft"] == pytest.approx(1311.6)
    # legacy SW Hydraulics columns joined by matched BRn name
    assert rec["gw_head_ft"] == pytest.approx(1311.5)
    assert rec["screen_elev_ft"] is None


def test_refresh_preserves_hand_edits_and_sites(workbook_env):
    extracted, out = workbook_env
    # hand edits: one SITES value cell + WELLS editable cells
    wb = openpyxl.load_workbook(out)
    ws = wb["SITES"]
    notes_c = _col(ws, "notes")
    ws.cell(row=4, column=notes_c, value="HAND EDIT")
    flow_cms_c = _col(ws, "flow_cms")
    assert str(ws.cell(row=4, column=flow_cms_c).value).startswith("=")
    ww = wb["WELLS"]
    ww.cell(row=2, column=_col(ww, "screen_elev_ft"), value=1300.0)
    ww.cell(row=2, column=_col(ww, "name"), value="Renamed")
    wb.save(out)

    # source data changes: new head value + a brand-new well appears
    (extracted / "AA00001.json").write_text(
        json.dumps(_site_json(obs_head=9999.0, extra_well=True)), encoding="utf-8")
    master.refresh_wells()

    wb2 = openpyxl.load_workbook(out)
    assert wb2.sheetnames == wb.sheetnames        # WELLS recreated at the same index
    ws2 = wb2["SITES"]
    assert ws2.cell(row=4, column=notes_c).value == "HAND EDIT"
    assert str(ws2.cell(row=4, column=flow_cms_c).value).startswith("=")
    ww2 = wb2["WELLS"]
    hdr = [c.value for c in ww2[1]]
    rows = [dict(zip(hdr, [c.value for c in r])) for r in ww2.iter_rows(min_row=2)]
    by_key = {(r["site_id"], r["obs_name"]): r for r in rows}
    r1 = by_key[("AA00001", "hed1")]
    assert r1["name"] == "Renamed"                       # hand edit preserved
    assert r1["screen_elev_ft"] == pytest.approx(1300.0)  # hand edit preserved
    assert r1["obs_head_ft"] == pytest.approx(9999.0)     # factory cell refreshed
    assert ("AA00001", "hed2") in by_key                  # new well appended


def test_read_wells_groups_by_site(workbook_env, monkeypatch):
    _, out = workbook_env
    got = master.read_wells()
    assert set(got) == {"AA00001"}
    assert got["AA00001"][0]["obs_name"] == "hed1"
    assert got["AA00001"][0]["lat"] == pytest.approx(30.0)


def test_read_wells_tolerates_missing_or_legacy(tmp_path, monkeypatch):
    out = tmp_path / "m.xlsx"
    monkeypatch.setattr(master, "OUT_XLSX", out)
    assert master.read_wells() == {}              # file missing
    wb = openpyxl.Workbook()
    wb.active.title = "SITES"
    ww = wb.create_sheet("WELLS")
    for j, h in enumerate(["site_id", "well", "sw_head_ft"], 1):   # legacy header
        ww.cell(row=1, column=j, value=h)
    ww.cell(row=2, column=1, value="AA00001")
    wb.save(out)
    assert master.read_wells() == {}              # legacy format ignored


def test_refresh_requires_existing_workbook(tmp_path, monkeypatch):
    monkeypatch.setattr(master, "OUT_XLSX", tmp_path / "missing.xlsx")
    monkeypatch.setattr(master, "EXTRACTED", tmp_path)
    with pytest.raises(SystemExit):
        master.refresh_wells()


def test_wells_rows_flag_head_conflicts():
    """A named legacy join whose gw head disagrees with the hob head gets a note
    (the LL01096 lesson: GMS csv names and SW Hydraulics head order conflict)."""
    site = _site_json()
    site["sw_hydraulics"]["wells"]["BR1"]["gw_head_ft"] = 1313.5   # 1.9 ft off
    rows = master._wells_rows_from_extracted([site])
    assert "review name and datum" in rows[0]["note"]
    # agreement within 0.2 ft stays quiet
    site["sw_hydraulics"]["wells"]["BR1"]["gw_head_ft"] = 1311.61
    rows = master._wells_rows_from_extracted([site])
    assert rows[0]["note"] is None
